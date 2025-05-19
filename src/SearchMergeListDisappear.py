import os
import shutil
import re
import logging
import csv
from datetime import datetime, timedelta
from docx import Document
from PyPDF2 import PdfReader
import openpyxl
from PyQt6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
    QFileDialog, QLineEdit, QProgressBar, QTableWidget, QTableWidgetItem,
    QHeaderView, QMessageBox, QCheckBox, QFrame, QComboBox, QSpinBox,
    QScrollArea, QSizePolicy, QTabWidget, QMenu
)
from PyQt6.QtGui import QIcon, QAction
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QTimer

logging.basicConfig(
    filename='file_search_errors.log',
    level=logging.ERROR,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# -------------------- Search Thread --------------------
class SearchThread(QThread):
    update_progress = pyqtSignal(int, str)
    search_complete = pyqtSignal(list)
    error_occurred = pyqtSignal(str)

    def __init__(self, search_params):
        super().__init__()
        self.search_params = search_params
        self.stop_search = False

    def run(self):
        try:
            source_loc = self.search_params['source_loc']
            out_loc = self.search_params['out_loc']
            search_string = self.search_params['search_string']
            case_sensitive = self.search_params['case_sensitive']
            whole_word = self.search_params['whole_word']
            use_regex = self.search_params['use_regex']
            file_types = self.search_params['file_types']

            will_save = bool(out_loc)
            if will_save:
                os.makedirs(out_loc, exist_ok=True)
            
            matching_files = []
            if os.path.isfile(source_loc):
                file_list = [source_loc] if source_loc.lower().endswith(tuple(file_types)) else []
            else:
                file_list = [
                    os.path.join(root, file) 
                    for root, _, files in os.walk(source_loc) 
                    for file in files 
                    if file.lower().endswith(tuple(file_types))
                ]

            total_files = len(file_list)
            self.update_progress.emit(0, f"Found {total_files} files to search")

            for i, file_path in enumerate(file_list):
                if self.stop_search:
                    break

                filename = os.path.basename(file_path)
                ext = os.path.splitext(filename)[1].lower()
                
                occurrences, locations = self.search_in_file(
                    file_path, ext, search_string, 
                    case_sensitive, whole_word, use_regex
                )

                if occurrences > 0:
                    if will_save:
                        try:
                            dest_path = os.path.join(out_loc, filename)
                            shutil.copy2(file_path, dest_path)
                            matching_files.append([filename, occurrences, locations, dest_path])
                        except Exception as e:
                            self.error_occurred.emit(f"Failed to copy {filename}: {str(e)}")
                    else:
                        matching_files.append([filename, occurrences, locations, "Not saved"])

                progress = int((i + 1) / total_files * 100)
                self.update_progress.emit(progress, f"Processing {filename}...")

            self.search_complete.emit(matching_files)

        except Exception as e:
            self.error_occurred.emit(f"Search error: {str(e)}")

    def search_in_file(self, file_path, extension, search_string, case_sensitive, whole_word, use_regex):
        try:
            occurrences, locations = 0, []
            flags = 0 if case_sensitive else re.IGNORECASE
            
            if use_regex:
                pattern = search_string
            else:
                pattern = re.escape(search_string)
                if whole_word:
                    pattern = r'\b' + pattern + r'\b'

            try:
                regex = re.compile(pattern, flags)
            except re.error as e:
                return 0, f"Invalid regex: {str(e)}"

            if extension == '.txt':
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    for i, line in enumerate(f, start=1):
                        matches = regex.finditer(line)
                        for match in matches:
                            occurrences += 1
                            locations.append(f"Line {i} (Pos {match.start()+1})")

            elif extension == '.docx':
                doc = Document(file_path)
                for i, para in enumerate(doc.paragraphs, start=1):
                    matches = regex.finditer(para.text)
                    for match in matches:
                        occurrences += 1
                        locations.append(f"Paragraph {i} (Pos {match.start()+1})")

            elif extension == '.xlsx':
                workbook = openpyxl.load_workbook(file_path)
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    for row in sheet.iter_rows():
                        for cell in row:
                            if cell.value:
                                matches = regex.finditer(str(cell.value))
                                for match in matches:
                                    occurrences += 1
                                    locations.append(f"Sheet '{sheet_name}' Cell {cell.coordinate}")

            elif extension == '.pdf':
                reader = PdfReader(file_path)
                for i, page in enumerate(reader.pages, start=1):
                    text = page.extract_text() or ""
                    matches = regex.finditer(text)
                    for match in matches:
                        occurrences += 1
                        locations.append(f"Page {i} (Pos {match.start()+1})")

            return occurrences, ', '.join(locations[:3]) + ('...' if len(locations) > 3 else '')
        
        except Exception as e:
            logging.error(f"Error processing {file_path}: {e}")
            return 0, "Error"

    def stop(self):
        self.stop_search = True
        self.update_progress.emit(0, "Search stopped")

# -------------------- CSV Merge Thread --------------------
class CSVThread(QThread):
    update_progress = pyqtSignal(int, str)
    merge_complete = pyqtSignal(str)
    error_occurred = pyqtSignal(str)

    def __init__(self, input_files, output_file, include_headers):
        super().__init__()
        self.input_files = input_files
        self.output_file = output_file
        self.include_headers = include_headers
        self.stop_merge = False

    def run(self):
        try:
            total_files = len(self.input_files)
            if total_files == 0:
                self.error_occurred.emit("No CSV files selected for merging")
                return

            self.update_progress.emit(0, "Starting CSV merge...")
            headers = []
            merged_data = []

            for i, file_path in enumerate(self.input_files):
                if self.stop_merge:
                    break

                self.update_progress.emit(int((i + 1) / total_files * 50),
                                        f"Processing {os.path.basename(file_path)}...")

                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    reader = csv.reader(f)
                    file_headers = next(reader)

                    if not headers:
                        headers = file_headers
                        if self.include_headers:
                            merged_data.append(headers)
                    elif file_headers != headers:
                        self.error_occurred.emit(
                            f"Header mismatch in {os.path.basename(file_path)}\n"
                            f"Expected: {headers}\nFound: {file_headers}"
                        )
                        return

                    for row in reader:
                        merged_data.append(row)

            self.update_progress.emit(75, "Writing merged file...")

            with open(self.output_file, 'w', encoding='utf-8', newline='') as f:
                writer = csv.writer(f)
                writer.writerows(merged_data)

            self.merge_complete.emit(f"Successfully merged {total_files} files into:\n{self.output_file}")

        except Exception as e:
            self.error_occurred.emit(f"CSV merge error: {str(e)}")

    def stop(self):
        self.stop_merge = True
        self.update_progress.emit(0, "CSV merge stopped")

# -------------------- Main Application --------------------
class FileSearchApp(QWidget):
    def __init__(self):
        super().__init__()
        self.search_thread = None
        self.csv_thread = None
        self.supported_file_types = ('.txt', '.docx', '.xlsx', '.pdf')
        self.file_timers = {}
        self.csv_files = []
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("File Search & CSV Merger")
        self.setWindowIcon(QIcon("icon.png" if os.path.exists("icon.png") else None))
        
        # Create tab widget
        tabs = QTabWidget()
        
        # Create search tab
        search_tab = QWidget()
        self.setup_search_tab(search_tab)
        
        # Create CSV merge tab
        csv_tab = QWidget()
        self.setup_csv_tab(csv_tab)
        
        # Add tabs
        tabs.addTab(search_tab, "File Search")
        tabs.addTab(csv_tab, "CSV Merger")
        
        # Main layout
        main_layout = QVBoxLayout(self)
        main_layout.addWidget(tabs)
        
        self.apply_dark_theme()
        self.showMaximized()

    def setup_search_tab(self, tab):
        # Main scroll area
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        
        # Container widget
        container = QWidget()
        container.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        layout = QVBoxLayout(container)
        layout.setContentsMargins(10, 10, 10, 10)
        
        # Input Section
        input_group = QFrame()
        input_group.setFrameShape(QFrame.Shape.StyledPanel)
        input_layout = QVBoxLayout()
        
        self.input_label = QLabel("Select Input Folder or File:")
        self.input_path = QLineEdit()
        self.input_path.setPlaceholderText("Path to search in")
        
        input_btn_layout = QHBoxLayout()
        self.input_button = QPushButton("Browse Folder")
        self.input_button.clicked.connect(self.select_input_folder)
        self.file_button = QPushButton("Browse File")
        self.file_button.clicked.connect(self.select_input_file)
        input_btn_layout.addWidget(self.input_button)
        input_btn_layout.addWidget(self.file_button)
        
        input_layout.addWidget(self.input_label)
        input_layout.addWidget(self.input_path)
        input_layout.addLayout(input_btn_layout)
        input_group.setLayout(input_layout)
        
        # Search Options
        options_group = QFrame()
        options_group.setFrameShape(QFrame.Shape.StyledPanel)
        options_layout = QVBoxLayout()
        
        self.search_label = QLabel("Search Text:")
        self.search_text = QLineEdit()
        self.search_text.setPlaceholderText("Text or pattern to search for")
        
        self.case_sensitive = QCheckBox("Case sensitive")
        self.whole_word = QCheckBox("Whole word only")
        self.use_regex = QCheckBox("Use regular expressions")
        
        file_type_layout = QHBoxLayout()
        file_type_layout.addWidget(QLabel("File Types:"))
        self.file_type_combo = QComboBox()
        self.file_type_combo.addItems([
            "All Supported (.txt, .docx, .xlsx, .pdf)", 
            "Text Files (.txt)", 
            "Word Documents (.docx)",
            "Excel Files (.xlsx)",
            "PDF Documents (.pdf)"
        ])
        file_type_layout.addWidget(self.file_type_combo)
        
        options_layout.addWidget(self.search_label)
        options_layout.addWidget(self.search_text)
        options_layout.addWidget(self.case_sensitive)
        options_layout.addWidget(self.whole_word)
        options_layout.addWidget(self.use_regex)
        options_layout.addLayout(file_type_layout)
        options_group.setLayout(options_layout)
        
        # Output Section
        output_group = QFrame()
        output_group.setFrameShape(QFrame.Shape.StyledPanel)
        output_layout = QVBoxLayout()
        
        self.output_label = QLabel("Output Folder (leave empty to not save copies):")
        self.output_path = QLineEdit()
        self.output_path.setPlaceholderText("Where to save matching files")
        self.output_button = QPushButton("Browse")
        self.output_button.clicked.connect(self.select_output_folder)
        
        # Expiration Settings
        self.expiration_check = QCheckBox("Enable file expiration")
        self.expiration_check.stateChanged.connect(self.toggle_expiration_controls)
        
        expiration_time_layout = QHBoxLayout()
        self.expiration_time_label = QLabel("Expire after:")
        self.expiration_time = QSpinBox()
        self.expiration_time.setRange(1, 86400)
        self.expiration_time.setValue(10)  # Default to 10 seconds for testing
        self.expiration_time_unit = QComboBox()
        self.expiration_time_unit.addItems(["seconds", "minutes", "hours", "days"])
        expiration_time_layout.addWidget(self.expiration_time_label)
        expiration_time_layout.addWidget(self.expiration_time)
        expiration_time_layout.addWidget(self.expiration_time_unit)
        
        expiration_folder_layout = QHBoxLayout()
        self.expiration_folder_label = QLabel("Move to:")
        self.expiration_folder = QLineEdit()
        self.expiration_folder.setPlaceholderText("Folder for expired files")
        self.expiration_folder_button = QPushButton("Browse")
        self.expiration_folder_button.clicked.connect(self.select_expiration_folder)
        expiration_folder_layout.addWidget(self.expiration_folder_label)
        expiration_folder_layout.addWidget(self.expiration_folder)
        expiration_folder_layout.addWidget(self.expiration_folder_button)
        
        output_layout.addWidget(self.output_label)
        output_layout.addWidget(self.output_path)
        output_layout.addWidget(self.output_button)
        output_layout.addWidget(self.expiration_check)
        output_layout.addLayout(expiration_time_layout)
        output_layout.addLayout(expiration_folder_layout)
        output_group.setLayout(output_layout)
        
        # Action Buttons
        action_layout = QHBoxLayout()
        self.search_button = QPushButton("Start Search")
        self.search_button.clicked.connect(self.start_search)
        self.stop_button = QPushButton("Stop")
        self.stop_button.clicked.connect(self.stop_searching)
        self.stop_button.setEnabled(False)
        self.export_button = QPushButton("Export Results")
        self.export_button.clicked.connect(self.export_results)
        action_layout.addWidget(self.search_button)
        action_layout.addWidget(self.stop_button)
        action_layout.addWidget(self.export_button)
        
        # Progress Bar
        self.progress = QProgressBar()
        self.progress.setTextVisible(True)
        
        # Status Bar
        self.status_bar = QLabel("Ready")
        self.status_bar.setAlignment(Qt.AlignmentFlag.AlignLeft)
        
        # Results Table
        self.result_table = QTableWidget()
        self.result_table.setColumnCount(4)
        self.result_table.setHorizontalHeaderLabels(["File Name", "Occurrences", "Locations", "Saved To"])
        self.result_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.result_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.result_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.result_table.setMinimumHeight(300)
        self.result_table.verticalHeader().setDefaultSectionSize(30)
        
        # Add widgets to layout
        layout.addWidget(input_group)
        layout.addWidget(options_group)
        layout.addWidget(output_group)
        layout.addLayout(action_layout)
        layout.addWidget(self.progress)
        layout.addWidget(self.result_table)
        layout.addWidget(self.status_bar)
        layout.addStretch()
        
        # Set up scroll area
        scroll.setWidget(container)
        
        # Tab layout
        tab_layout = QVBoxLayout(tab)
        tab_layout.addWidget(scroll)

    def setup_csv_tab(self, tab):
        layout = QVBoxLayout(tab)
        
        # CSV Merge Group
        csv_merge_group = QFrame()
        csv_merge_group.setFrameShape(QFrame.Shape.StyledPanel)
        csv_merge_layout = QVBoxLayout()
        
        csv_merge_label = QLabel("CSV File Operations:")
        self.csv_files_path = QLineEdit()
        self.csv_files_path.setReadOnly(True)
        self.csv_files_path.setPlaceholderText("Selected CSV files will appear here")

        csv_merge_btn_layout = QHBoxLayout()
        csv_select_button = QPushButton("Select CSV Files")
        csv_select_button.clicked.connect(self.select_csv_files)
        csv_clear_button = QPushButton("Clear Selection")
        csv_clear_button.clicked.connect(self.clear_csv_selection)
        csv_merge_btn_layout.addWidget(csv_select_button)
        csv_merge_btn_layout.addWidget(csv_clear_button)

        csv_merge_options = QHBoxLayout()
        self.csv_include_headers = QCheckBox("Include headers in output")
        self.csv_include_headers.setChecked(True)
        csv_merge_options.addWidget(self.csv_include_headers)

        csv_merge_action_layout = QHBoxLayout()
        self.csv_merge_button = QPushButton("Merge CSV Files")
        self.csv_merge_button.clicked.connect(self.start_csv_merge)
        self.csv_stop_merge_button = QPushButton("Stop Merge")
        self.csv_stop_merge_button.clicked.connect(self.stop_csv_merge)
        self.csv_stop_merge_button.setEnabled(False)
        csv_merge_action_layout.addWidget(self.csv_merge_button)
        csv_merge_action_layout.addWidget(self.csv_stop_merge_button)

        csv_merge_layout.addWidget(csv_merge_label)
        csv_merge_layout.addWidget(self.csv_files_path)
        csv_merge_layout.addLayout(csv_merge_btn_layout)
        csv_merge_layout.addLayout(csv_merge_options)
        csv_merge_layout.addLayout(csv_merge_action_layout)
        csv_merge_group.setLayout(csv_merge_layout)

        # CSV Progress and Status
        self.csv_progress = QProgressBar()
        self.csv_progress.setValue(0)
        self.csv_status_bar = QLabel("Ready")
        self.csv_status_bar.setAlignment(Qt.AlignmentFlag.AlignLeft)

        layout.addWidget(csv_merge_group)
        layout.addWidget(self.csv_progress)
        layout.addWidget(self.csv_status_bar)
        layout.addStretch()

    # -------------------- File Search Methods --------------------
    def toggle_expiration_controls(self, state):
        enabled = state == Qt.CheckState.Checked.value
        self.expiration_time_label.setEnabled(enabled)
        self.expiration_time.setEnabled(enabled)
        self.expiration_time_unit.setEnabled(enabled)
        self.expiration_folder_label.setEnabled(enabled)
        self.expiration_folder.setEnabled(enabled)
        self.expiration_folder_button.setEnabled(enabled)

    def select_input_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Input Folder")
        if folder:
            self.input_path.setText(folder)

    def select_input_file(self):
        file, _ = QFileDialog.getOpenFileName(
            self, "Select File", "", 
            "Supported Files (*.txt *.docx *.xlsx *.pdf);;All Files (*)"
        )
        if file:
            self.input_path.setText(file)

    def select_output_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Output Folder")
        if folder:
            self.output_path.setText(folder)

    def select_expiration_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Expiration Folder")
        if folder:
            self.expiration_folder.setText(folder)
            os.makedirs(folder, exist_ok=True)

    def start_search(self):
        source_loc = self.input_path.text()
        out_loc = self.output_path.text()
        search_string = self.search_text.text()

        if not source_loc:
            QMessageBox.warning(self, "Missing Input", "Please select input location!")
            return

        if not search_string:
            QMessageBox.warning(self, "Missing Input", "Please enter search text!")
            return

        selected_type = self.file_type_combo.currentText()
        if selected_type == "All Supported (.txt, .docx, .xlsx, .pdf)":
            file_types = self.supported_file_types
        elif selected_type == "Text Files (.txt)":
            file_types = ('.txt',)
        elif selected_type == "Word Documents (.docx)":
            file_types = ('.docx',)
        elif selected_type == "Excel Files (.xlsx)":
            file_types = ('.xlsx',)
        elif selected_type == "PDF Documents (.pdf)":
            file_types = ('.pdf',)

        search_params = {
            'source_loc': source_loc,
            'out_loc': out_loc,
            'search_string': search_string,
            'case_sensitive': self.case_sensitive.isChecked(),
            'whole_word': self.whole_word.isChecked(),
            'use_regex': self.use_regex.isChecked(),
            'file_types': file_types
        }

        self.result_table.setRowCount(0)
        self.status_bar.setText("Starting search...")

        self.search_thread = SearchThread(search_params)
        self.search_thread.update_progress.connect(self.update_progress_status)
        self.search_thread.search_complete.connect(self.search_completed)
        self.search_thread.error_occurred.connect(self.handle_error)
        
        self.search_button.setEnabled(False)
        self.stop_button.setEnabled(True)
        self.search_thread.start()

    def stop_searching(self):
        if self.search_thread and self.search_thread.isRunning():
            self.search_thread.stop()
            self.status_bar.setText("Search stopped")

    def update_progress_status(self, value, message):
        self.progress.setValue(value)
        self.status_bar.setText(message)

    def search_completed(self, results):
        self.search_button.setEnabled(True)
        self.stop_button.setEnabled(False)
        self.progress.setValue(100)
        
        if not results:
            self.status_bar.setText("Search completed - no matches found")
            QMessageBox.information(self, "Search Complete", "No matching files found.")
            return

        self.update_table(results)
        self.status_bar.setText(f"Search completed - {len(results)} matches found")

        # Schedule expiration for saved files if enabled
        if self.expiration_check.isChecked() and self.output_path.text() and self.expiration_folder.text():
            for row_data in results:
                if row_data[3] != "Not saved":
                    file_path = row_data[3]
                    self.schedule_file_expiration(file_path)

    def schedule_file_expiration(self, file_path):
        """Reliable file expiration scheduling"""
        if not (self.expiration_check.isChecked() and self.expiration_folder.text()):
            return
            
        try:
            # Calculate time in seconds
            time_sec = self.expiration_time.value()
            unit = self.expiration_time_unit.currentText()
            
            if unit == "minutes":
                time_sec *= 60
            elif unit == "hours":
                time_sec *= 3600
            elif unit == "days":
                time_sec *= 86400
            
            # Create and start timer (parented to self)
            timer = QTimer(self)
            timer.setSingleShot(True)
            
            # Use lambda with default argument to capture current file_path
            timer.timeout.connect(lambda f=file_path: self.move_expired_file(f))
            timer.start(time_sec * 1000)  # Convert to milliseconds
            
            # Store reference
            self.file_timers[file_path] = timer
            
            self.status_bar.setText(f"Scheduled to move {os.path.basename(file_path)} in {time_sec} seconds")
            
        except Exception as e:
            logging.error(f"Error scheduling expiration: {e}")
            self.status_bar.setText(f"Error scheduling expiration: {str(e)}")

    def move_expired_file(self, file_path):
        """Handle the actual file movement"""
        try:
            expiration_folder = self.expiration_folder.text()
            if not expiration_folder:
                return
                
            if not os.path.exists(file_path):
                self.update_file_status(file_path, "File missing")
                return
                
            dest_path = os.path.join(expiration_folder, os.path.basename(file_path))
            
            # Ensure destination exists
            os.makedirs(expiration_folder, exist_ok=True)
            
            # Perform the move
            shutil.move(file_path, dest_path)
            
            # Update UI
            self.update_file_status(file_path, f"Moved to {expiration_folder}")
            self.status_bar.setText(f"Moved {os.path.basename(file_path)} to expiration folder")
            
            # Clean up timer
            if file_path in self.file_timers:
                del self.file_timers[file_path]
                
        except Exception as e:
            error_msg = f"Failed to move {os.path.basename(file_path)}: {str(e)}"
            logging.error(error_msg)
            self.update_file_status(file_path, "Move failed")
            self.status_bar.setText(error_msg)

    def update_file_status(self, file_path, status):
        """Update status in results table"""
        filename = os.path.basename(file_path)
        for row in range(self.result_table.rowCount()):
            if self.result_table.item(row, 0).text() == filename:
                self.result_table.item(row, 3).setText(status)
                self.result_table.scrollToItem(self.result_table.item(row, 0))
                break

    def handle_error(self, error_msg):
        self.status_bar.setText(f"Error: {error_msg}")
        QMessageBox.critical(self, "Error", error_msg)
        self.search_button.setEnabled(True)
        self.stop_button.setEnabled(False)

    def update_table(self, file_list):
        self.result_table.setRowCount(len(file_list))
        for row, data in enumerate(file_list):
            for col, value in enumerate(data):
                item = QTableWidgetItem(str(value))
                if col == 1:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.result_table.setItem(row, col, item)
        
        self.result_table.resizeRowsToContents()

    def export_results(self):
        if self.result_table.rowCount() == 0:
            QMessageBox.warning(self, "No Results", "Nothing to export - no search results available")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "Save Results", "", 
            "CSV Files (*.csv);;Text Files (*.txt)"
        )

        if not path:
            return

        try:
            with open(path, 'w', encoding='utf-8') as f:
                headers = []
                for col in range(self.result_table.columnCount()):
                    headers.append(self.result_table.horizontalHeaderItem(col).text())
                f.write(','.join(headers) + '\n')

                for row in range(self.result_table.rowCount()):
                    row_data = []
                    for col in range(self.result_table.columnCount()):
                        item = self.result_table.item(row, col)
                        row_data.append(item.text() if item else '')
                    f.write(','.join(row_data) + '\n')

            QMessageBox.information(self, "Export Complete", f"Results exported to:\n{path}")

        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export results:\n{str(e)}")

    # -------------------- CSV Merge Methods --------------------
    def select_csv_files(self):
        files, _ = QFileDialog.getOpenFileNames(
            self, "Select CSV Files", "",
            "CSV Files (*.csv);;All Files (*)"
        )
        if files:
            self.csv_files = files
            self.csv_files_path.setText(f"{len(files)} files selected")
            self.csv_status_bar.setText(f"Selected {len(files)} CSV files")

    def clear_csv_selection(self):
        self.csv_files = []
        self.csv_files_path.clear()
        self.csv_files_path.setPlaceholderText("Selected CSV files will appear here")
        self.csv_status_bar.setText("Cleared CSV file selection")

    def start_csv_merge(self):
        if not self.csv_files:
            QMessageBox.warning(self, "No Files", "Please select CSV files to merge first!")
            return

        output_file, _ = QFileDialog.getSaveFileName(
            self, "Save Merged CSV File", "",
            "CSV Files (*.csv);;All Files (*)"
        )
        if not output_file:
            return

        if not output_file.lower().endswith('.csv'):
            output_file += '.csv'

        self.csv_thread = CSVThread(
            self.csv_files,
            output_file,
            self.csv_include_headers.isChecked()
        )

        self.csv_thread.update_progress.connect(self.update_csv_progress_status)
        self.csv_thread.merge_complete.connect(self.csv_merge_completed)
        self.csv_thread.error_occurred.connect(self.handle_csv_error)

        self.csv_thread.start()

        self.csv_status_bar.setText("Starting CSV merge...")
        self.csv_merge_button.setEnabled(False)
        self.csv_stop_merge_button.setEnabled(True)

    def stop_csv_merge(self):
        if self.csv_thread and self.csv_thread.isRunning():
            self.csv_thread.stop()
            self.csv_status_bar.setText("CSV merge stopped")
            self.csv_merge_button.setEnabled(True)
            self.csv_stop_merge_button.setEnabled(False)

    def csv_merge_completed(self, message):
        self.csv_merge_button.setEnabled(True)
        self.csv_stop_merge_button.setEnabled(False)
        self.csv_status_bar.setText(message)
        QMessageBox.information(self, "Merge Complete", message)

    def update_csv_progress_status(self, value, message):
        self.csv_progress.setValue(value)
        self.csv_status_bar.setText(message)

    def handle_csv_error(self, message):
        self.csv_merge_button.setEnabled(True)
        self.csv_stop_merge_button.setEnabled(False)
        self.csv_status_bar.setText(message)
        QMessageBox.critical(self, "Error", message)

    # -------------------- Common Methods --------------------
    def apply_dark_theme(self):
        self.setStyleSheet("""
            QWidget {
                background-color: #2b2b2b;
                color: #e0e0e0;
                font-family: Segoe UI;
                font-size: 12px;
                border: none;
            }
            QPushButton {
                background-color: #3c3c3c;
                border: 1px solid #555;
                border-radius: 4px;
                padding: 5px 10px;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #4a4a4a;
            }
            QPushButton:pressed {
                background-color: #2a2a2a;
            }
            QLineEdit, QTableWidget, QComboBox, QSpinBox {
                background-color: #3c3c3c;
                border: 1px solid #555;
                border-radius: 4px;
                padding: 5px;
            }
            QProgressBar {
                border: 1px solid #555;
                border-radius: 4px;
                text-align: center;
                height: 20px;
            }
            QProgressBar::chunk {
                background-color: #4CAF50;
                width: 10px;
            }
            QHeaderView::section {
                background-color: #3c3c3c;
                padding: 5px;
                border: none;
            }
            QFrame {
                border-radius: 4px;
            }
            QCheckBox {
                spacing: 5px;
            }
            QScrollArea {
                border: none;
            }
            QScrollBar:vertical {
                width: 12px;
                background: #2b2b2b;
            }
            QScrollBar::handle:vertical {
                background: #4a4a4a;
                min-height: 20px;
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                background: none;
            }
            QTableWidget {
                font-size: 13px;
            }
            QTableWidget::item {
                padding: 5px;
            }
            QTabWidget::pane {
                border: 1px solid #444;
                top: -1px;
            }
            QTabBar::tab {
                background: #3c3c3c;
                border: 1px solid #444;
                padding: 5px 10px;
            }
            QTabBar::tab:selected {
                background: #4a4a4a;
                border-bottom-color: #4CAF50;
            }
        """)

    def closeEvent(self, event):
        # Clean up all timers
        for timer in self.file_timers.values():
            timer.stop()
            
        if self.search_thread and self.search_thread.isRunning():
            self.search_thread.stop()
            self.search_thread.wait(2000)
            
        if self.csv_thread and self.csv_thread.isRunning():
            self.csv_thread.stop()
            self.csv_thread.wait(2000)
            
        event.accept()

if __name__ == "__main__":
    import sys
    app = QApplication(sys.argv)
    window = FileSearchApp()
    window.show()
    sys.exit(app.exec())
